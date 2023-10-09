from datetime import datetime, timedelta

from django.db.models import Min, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from products.models import Product
from rest_framework import decorators, mixins, response, status, viewsets
from sales.models import Forecast, Sale
from stores.models import Store
from users.models import User

from . import serializers


class UserViewSet(mixins.RetrieveModelMixin,
                  viewsets.GenericViewSet):
    '''Вьюсет для модели User.'''

    queryset = User.objects.all()
    serializer_class = serializers.UserSerializer


class StoreViewSet(mixins.ListModelMixin,
                   viewsets.GenericViewSet):
    '''Вьюсет для модели Store.'''

    queryset = Store.objects.all()
    serializer_class = serializers.StoreSerializer


class SaleViewSet(mixins.ListModelMixin,
                  viewsets.GenericViewSet):
    '''Вьюсет для модели Sale.'''

    queryset = Sale.objects.all()
    serializer_class = serializers.SaleSerializer

    def get_sale_data(self, sku, start_date, end_date, is_rub_included=False):
        '''Метод для получения сериализованных данных о продажах.'''
        if start_date:
            query = Sale.objects.filter(
                sku=sku,
                date__range=(start_date, end_date)
                ).values('date').annotate(
                    total_sales_units=Sum('sales_units')
                )
            if is_rub_included:
                query = query.annotate(total_sales_rub=Sum('sales_rub'))

            return serializers.FactSerializer(query, many=True)

        return []

    @decorators.action(methods=['GET'], detail=False, url_path='21-days')
    def sales_21_days(self, request):
        '''Получение данных о продажах по паре товар-магазин за 21 день.'''
        store = get_object_or_404(
            Store, store=request.query_params.get('store')
        ).store
        sku = get_object_or_404(
            Product, sku=request.query_params.get('sku')
        ).sku

        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=21)

        serialized_data = self.get_sale_data(sku, start_date, end_date, False)

        response_data = {
            "store": store,
            "sku": sku,
            "fact": serialized_data.data
        }

        return response.Response({"data": response_data},
                                 status=status.HTTP_200_OK)

    @decorators.action(methods=['GET'], detail=False, url_path='total')
    def sales_total(self, request):
        '''Получение данных о продажах товара по всем магазинам
        в штуках и рублях за весь период.'''
        sku = get_object_or_404(
            Product, sku=request.query_params.get('sku')
            ).sku
        start_date = Sale.objects.aggregate(min_date=Min('date'))['min_date']
        end_date = request.query_params.get('end_date')

        end_date = (datetime.strptime(end_date, '%Y-%m-%d').date() if end_date
                    else datetime.now().date())

        serialized_data = self.get_sale_data(sku, start_date, end_date, True)

        response_data = {
            "sku": sku,
            "fact": serialized_data.data
        }

        return response.Response({"data": response_data},
                                 status=status.HTTP_200_OK)


class ForecastViewSet(mixins.ListModelMixin,
                      mixins.CreateModelMixin,
                      viewsets.GenericViewSet):
    '''Вьюсет для модели Forecast.'''

    queryset = Forecast.objects.all()
    serializer_class = serializers.ForecastSerializer

    @decorators.action(methods=['GET'], detail=False, url_path='download')
    def download_xlsx(self, request):
        '''Метод для скачивания xlsx-файла с прогнозными данными продаж.'''
        store = get_object_or_404(
            Store, store=request.query_params.get('store')
        ).store
        sku = get_object_or_404(
            Product, sku=request.query_params.get('sku')
        ).sku
        forecast_date = get_object_or_404(
            Forecast, forecast_date=request.query_params.get('forecast_date')
        ).forecast_date

        forecast = Forecast.objects.filter(
            store=store,
            sku=sku
        ).values('forecast')

        serializer = serializers.DownloadForecastSerializer(forecast,
                                                            many=True)

        forecast = {key: value for sublist in serializer.data
                    for key, value in sublist['forecast'].items()}

        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Дата'
        ws['B1'] = 'Продажи, шт.'

        row_num = 2
        for date, items in forecast.items():
            ws.cell(row=row_num, column=1, value=date)
            ws.cell(row=row_num, column=2, value=items)
            row_num += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-'
                                'officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="sales_{sku}_{forecast_date}.xlsx"'
        )

        wb.save(response)

        return response


class ProductViewSet(mixins.ListModelMixin,
                     viewsets.GenericViewSet):
    '''Вьюсет для модели Product.'''

    queryset = Product.objects.all()
    serializer_class = serializers.ProductSerializer
